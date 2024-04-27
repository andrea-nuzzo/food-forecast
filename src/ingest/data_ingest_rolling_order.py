import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter, column_index_from_string
from copy import copy
import re
from datetime import datetime
import os
import logging
from utils.logging import setup_logging
from transform.data_transform import DataTransform

setup_logging()


class DataIngestRollingOrder:
    def __init__(self, data, start_date,end_date):
        self.data = data
        self.start_date = start_date
        self.end_date = end_date
        self.stock_date = None
        self.file_path = None

    def finding_rolling_file(self):
        company_name = self.data["company_name"].iloc[0]
        company_name = company_name.lower()
        folder_path = "../data/rolling_analysis_order/"
        files = os.listdir(folder_path)

        matching_files = [f for f in files if company_name in f.lower()]

        if len(matching_files) > 1:
            raise ValueError(
                f"More files found for the company {company_name}: {matching_files}"
            )

        elif matching_files:
            self.file_path = os.path.join(folder_path, matching_files[0])
            data_rolling = pd.read_excel(self.file_path)
           
            return data_rolling
        else:
            print(f"No files found for the company {company_name}")
            return None
    
    def handle_dates(self, date_input):
        try:
            if not isinstance(date_input, str):
                date_input = str(date_input)
            
            date_parts = date_input.split('.')
            if len(date_parts) == 2:
                date_input += f".{datetime.now().year}"
            
            formatted_date = datetime.strptime(date_input, "%d.%m.%Y")
        except ValueError as e:
            logging.error(f"Invalid date format: {date_input} - Error: {str(e)}")
            raise
        except Exception as e:
            logging.error("Error ind ate formatting process: " + str(e))
            raise
        
        return formatted_date
    
    def get_stock(self, data):
        try:
            raw_stock_date = data.iat[4, -1]
            self.stock_date = self.handle_dates(raw_stock_date)
            
            if isinstance(self.start_date, str):
                self.start_date = datetime.strptime(self.start_date, "%d/%m/%Y")
            elif not isinstance(self.start_date, datetime):
                raise TypeError("start_date must be a string or a datetime object.")

            if self.stock_date.strftime("%d/%m/%Y") == self.start_date.strftime("%d/%m/%Y"):
                num_columns = data.shape[1]
                stock_data = data.iloc[5:85, [0, 1, 2, num_columns - 1]]
                stock_data = stock_data.dropna()
                stock_data = stock_data.rename(
                    columns={
                        stock_data.columns[0]: "product__external_id",
                        stock_data.columns[1]: "einheit",
                        stock_data.columns[2]: "name",
                        stock_data.columns[3]: "bestand",
                    })
                return stock_data
            else:
                logging.error("Dates do not match.")
                raise ValueError("Stock date and start date do not match.")
        except Exception as e:
            logging.error("General error: " + str(e))
            raise
    
    def update_stock(self, data, stock_data):
        merged_df = pd.merge(stock_data, data, on='product__external_id', how='left')
        merged_df['product_quantity'] = merged_df['product_quantity'].fillna(0)
        merged_df['neustand'] = merged_df['bestand'] - merged_df['product_quantity']
        merged_df.insert(0, 'bestellt', 0)
        finale_df = merged_df[['bestellt','product_quantity', 'neustand']]
        
        return finale_df

    def update_file(self,ingest_data):
        wb = openpyxl.load_workbook(self.file_path)
        sheet = wb['Bestellubersicht']
        max_column = sheet.max_column
        
        if isinstance(self.start_date, str):
            self.start_date = datetime.strptime(self.start_date, '%d/%m/%Y')
        if isinstance(self.end_date, str):
            self.end_date = datetime.strptime(self.end_date, '%d/%m/%Y')
        
        for col_offset in range(4):
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                original_cell = row[max_column - 3 + col_offset - 1]
                new_cell_column = max_column + col_offset + 1
                new_cell = sheet.cell(row=original_cell.row, column=new_cell_column)
                # Copia il contenuto e lo stile
                if original_cell.has_style:
                    new_cell.font = copy(original_cell.font)
                    new_cell.border = copy(original_cell.border)
                    new_cell.fill = copy(original_cell.fill)
                    new_cell.number_format = copy(original_cell.number_format)
                    new_cell.protection = copy(original_cell.protection)
                    new_cell.alignment = copy(original_cell.alignment)
                # Aggiorna e copia la formula
                if isinstance(original_cell.value, str) and original_cell.value.startswith('='):
                    new_cell.value = self.adjust_formula(original_cell.value, 4)
                # Altrimenti copia il valore
                else:
                    new_cell.value = original_cell.value
        
        start_column = sheet.max_column - 2

        start_date_formatted = self.start_date.strftime('%d/%m/%Y')
        end_date_formatted = self.end_date.strftime('%d/%m/%Y')
        
        # Imposta le date nelle celle specificate
        sheet.cell(row=6, column=start_column).value = start_date_formatted
        sheet.cell(row=6, column=start_column + 1).value = f'{start_date_formatted} - {end_date_formatted}'
        sheet.cell(row=6, column=start_column + 2).value = end_date_formatted
        
        
        for r_idx, row in enumerate(dataframe_to_rows(ingest_data, index=False, header=False), 7):
            for c_idx, value in enumerate(row[:2], start_column):  # Parti dalla terzultima colonna
                sheet.cell(row=r_idx, column=c_idx, value=value)

        # Salva il nuovo file Excel
        # Ottieni il timestamp corrente
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Costruisci il nuovo percorso del file
        base_folder = os.path.dirname(self.file_path)
        parent_folder = os.path.dirname(base_folder)
        processed_folder = os.path.join(parent_folder, 'processed')
        
        original_filename = os.path.basename(self.file_path)  # Ottiene il nome del file dall'intero percorso
        filename_without_extension = original_filename.replace('.xlsx', '')  # Rimuove l'estensione
        new_filename = f'{filename_without_extension}_{timestamp}.xlsx'  # Aggiunge il timestamp
        processed_folder = os.path.join(parent_folder, 'processed')
        
        # Assicurati che la cartella 'processed' esista, altrimenti creala
        if not os.path.exists(processed_folder):
            os.makedirs(processed_folder)
        
        # Costruisci il percorso completo per il nuovo file
        new_excel_path = os.path.join(processed_folder, new_filename)
        
        # Salva il workbook nel nuovo percorso
        wb.save(new_excel_path)
        print(f"File saved to {new_excel_path}")
        
    def adjust_formula(self, formula, offset):
        # Funzione per sostituire ciascun riferimento alla colonna nella formula
        def replace_cell(match):
            start_col = match.group(1)
            start_row = match.group(2)
            end_col = match.group(3)
            end_row = match.group(4)
            
            # Calcola la nuova colonna iniziando dall'offset dato
            new_start_col = get_column_letter(column_index_from_string(start_col) + offset)
            new_start_ref = f"{new_start_col}{start_row}"

            if end_col:  # Se è un range
                new_end_col = get_column_letter(column_index_from_string(end_col) + offset)
                new_end_ref = f"{new_end_col}{end_row}"
                return f"{new_start_ref}:{new_end_ref}"
            else:
                return new_start_ref

        # Espressione regolare per trovare riferimenti a singole celle e range di celle
        cell_pattern = re.compile(r'([A-Z]+)(\d+)(?::([A-Z]+)(\d+))?')
        new_formula = cell_pattern.sub(replace_cell, formula)
        return new_formula
    
    def ingest_data(self, data):

        data_rolling = self.finding_rolling_file()
        stock_data = self.get_stock(data_rolling)
        ingest_data = self.update_stock(data, stock_data)
        self.update_file(ingest_data)


        
